#!/usr/bin/python

###############################################################
#              ECE415 High Performance Computing              #
###############################################################
#                            Lab 1                            #
#                                                             #
# Project: Optimizing the performance of a grayscale Sobel    #
#          filter                                             #
#                                                             #
# File Description: Python script to automatically create     #
#                   an Excel spreadsheet containing all data  #
#                   across all experiments run on the Sobel   #
#                   filter optimization process. The script   #
#                   takes as input all output .runlog files   #
#                   which contain data in CSV format          #
#                                                             #
# Authors: Kyritsis Spyridon     AEM: 2697                    #
#          Karaiskos Charalampos AEM: 2765                    #
###############################################################

import os
import sys
import datetime
# import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


## *** Classes *** ##

## ** Legend ** ##
## Class containing the different legends used for easy lookup                ##
## The testnumarray will contain the names of the different input             ##
## executable files as they appear in the source/ directory                   ##
## The testcollumnindex will consist of the full name of the test as found in ##
## the output/ directory and will contain the cell position where information ##
## should be inserted for the current test                                    ##
class Legend:
  testname = []
  testnum = []


class messagecolours:
  INFO = "\033[94m"
  SUCCESS = "\033[92m"
  FAIL = '\033[91m'
  WARNING = '\033[93m'
  BOLD = '\033[2m'
  DEFAULT = '\033[0m'


class testtree:
  compilersarray = []
  compilerstartingcol = []

  optimizationsarray = [[]]
  optimizationsstartingcol = [[]]


sourcefilearray = []
sourcefilesnum = 0


## *** Globals *** ##
compilersnum = 0
testsnum = 0


## *** Functions *** ##

def xlref(row, column, zero_indexed = True):
  if zero_indexed:
    row += 1
    column += 1

  return get_column_letter(column) + str(row)


## ** print_header ** ##
## print a header message with basic information about the script ##
def print_header():

  os.system("clear")

  print(messagecolours.BOLD + messagecolours.INFO)
  print("Tool: CSV to Excel script")
  print("Date: " + datetime.datetime.now().strftime("%B %d, %Y %H:%M:%S"))
  print("Author: Spyros Kyritsis")
  print("Short description: Automatically create an Excel spreadsheet containing results from the Sobel filter optimization tests")

  print(messagecolours.DEFAULT)


## ** count_executables ** ##
## Walks through the source/ directory counting the executables ##
## and creating a legend with name-number correspondence        ##
def count_executables(sourcedir):

  testarray = Legend()
  i = 0

  for dirpath, subdirnames, files in os.walk(sourcedir):

    for name in sorted(subdirnames):
      # print(name + ", " + name.split('_')[0])

      testarray.testnum.append(int(name.split('_')[0]))
      testarray.testname.append(name.split('_', 1)[1])

      # print(str(testarray.testnum[i]) + ", " + testarray.testname[i])

      i = i + 1

  return testarray, i


## ** create_exceltree ** ##
def create_exceltree(outputdir):

  exceltree = testtree()

  ## Counter to iterate through the file list ##
  i = 0

  print(messagecolours.INFO + "INFO: Counting the different compilers and optimization levels..." + messagecolours.DEFAULT)

  ## Count compilers ##
  for dirpath, subdirnames, files in os.walk(outputdir):

    ## Get the compilers name ##
    for name in sorted(subdirnames):

      # print(name)

      compilername = name.rsplit('_', 1)[1].split('-')[0]

      # print(compilername)

      ## Add the compiler to the compiler array if it does not yet exist there ##
      if compilername not in exceltree.compilersarray:
        exceltree.compilersarray.append(compilername)
        exceltree.optimizationsarray.append([])
        exceltree.optimizationsstartingcol.append([])


      # print(exceltree.compilersarray)

      ## Get the compiler's index ##
      compilerindex = exceltree.compilersarray.index(compilername)

      # print(str(compilerindex))

      ## Split the line on the first '-' encountered and keep the right part ##
      ## Get the optimization flags ##
      optimizationname = name.split('-', 1)[1]

      # print(optimizationname)
      # print(len(exceltree.optimizationsarray[compilerindex]))

      ## If the optimization name does not exist in the optimization list of the current compiler, append it ##
      if not any(optimizationname in x for x in exceltree.optimizationsarray[compilerindex]):
        exceltree.optimizationsarray[compilerindex].append(optimizationname)

      # print(exceltree.optimizationsarray[compilerindex])
      # print(len(exceltree.optimizationsarray[compilerindex]))

  print(messagecolours.INFO + "INFO: Done!" + messagecolours.DEFAULT)

  ## Reset the loop variable ##
  i = 0

  # print("Compilersarray length: " + str(len(exceltree.compilersarray)))

  print(messagecolours.INFO + "INFO: Finding the starting column for each test..." + messagecolours.DEFAULT)

  ## Find all the starting columns ##
  ## compilerstartingcol[i] = compilerstartingcol[i] + len(optimizationarray[i]) * (totaltestsnum * 2) ##
  for i in range(0, len(exceltree.compilersarray)):

    # print("i = " + str(i))

    ## Starting column for each compiler cell ##
    if (i == 0):
      ## The first column of the spreadsheet for writing is col E ##
      exceltree.compilerstartingcol.append(datastartcol)

    else:
      exceltree.compilerstartingcol.append(exceltree.compilerstartingcol[i - 1] + len(exceltree.optimizationsarray[i - 1]) * (totaltestsnum * 2))

    ## Loop all optimizations of a compiler and calculate their starting column ##
    ## startingcol[i][j] = startingcol[i][j] + (totaltestsnum * 2)                   ##
    for j in range(0, len(exceltree.optimizationsarray[i])):

      # print("j = " + str(j))

      ## The first optimization column starts at the same column as the compiler ##
      if (j == 0):
        exceltree.optimizationsstartingcol[i].append(exceltree.compilerstartingcol[i])

      else:
        exceltree.optimizationsstartingcol[i].append(exceltree.optimizationsstartingcol[i][j - 1] + (totaltestsnum * 2))

  print(messagecolours.INFO + "INFO: Done!" + messagecolours.DEFAULT)

  return exceltree


## ** dump_exceltree_info ** ##
def dump_exceltree_info(exceltree):

  print(messagecolours.INFO + "INFO: The test structure is:" + messagecolours.DEFAULT)

  for i in range(0, len(exceltree.compilersarray)):
    print("Compiler name:" + str(exceltree.compilersarray[i]))
    print('\t' + "Optimization list: { " + str(exceltree.optimizationsarray[i]) + "}")


## ** build_excel_spreadsheet ** ##
def build_excel_spreadsheet(exceltree, builddir, outputdir):

  print(messagecolours.INFO + "INFO: Appending all CSV-formated output log files to the excel spreadsheet..." + messagecolours.DEFAULT)

  ## Create a writer for the above created Excel spreadsheet in order to append data to it ##
  writer = pd.ExcelWriter('results.xlsx', engine='xlsxwriter')

  for dirpath, subdirnames, files in os.walk(outputdir):

    for name in sorted(files):
      if not name.endswith(".runlog"):
        continue

      # print(name)

      dataframe = pd.read_csv(dirpath + '/' + name, header=0)

      # print(dataframe)

      ## Drop the min and max values from the CSV file to ensure ##
      ## results are not skewed due to random OS related events  ##
      max_index = dataframe[["Execution time"]].idxmax()
      min_index = dataframe[["Execution time"]].idxmin()

      # print(max_index)

      dataframe = dataframe.drop(max_index)
      dataframe = dataframe.drop(min_index)

      ## Get test number ##
      testnum = name.split('_', 1)[0]

      compilername = name.rsplit('_', 1)[1].split('-')[0]

      compilerindex = exceltree.compilersarray.index(compilername)

      # print(compilername)
      # print(str(compilerindex))

      ## Get the optimization flags ##
      ## Split the line on the first '-' encountered and keep the right part ##
      ## Then remove the .runlog suffix                                      ##
      optimizationname = name.split('-', 1)[1].split('.', 1)[0]

      optimizationindex = exceltree.optimizationsarray[compilerindex].index(optimizationname)

      startingcol = exceltree.compilerstartingcol[compilerindex] + (optimizationindex * totaltestsnum * 2) + (int(testnum) * 2) - 1

      # print("Starting col = " + str(startingcol))

      dataframe.to_excel(writer, sheet_name='Sheet1', startcol=startingcol, startrow=(datastartrow + 2), index=False)

  writer.save()

  print(messagecolours.INFO + "INFO: Done!" + messagecolours.DEFAULT)

  ## Create the spreadsheet header ##
  file_path = 'results.xlsx'
  wb = load_workbook(file_path)
  sheet1 = wb['Sheet1']  # or wb.active

  # print(len(exceltree.compilersarray))

  print(messagecolours.INFO + "INFO: Writing spreadsheet headers and footers..." + messagecolours.DEFAULT)

  ## Write compiler headers ##
  for i in range(0, len(exceltree.compilersarray)):

    # print("totaltestsnum = " + str(totaltestsnum) + " optimizationsarray[i] = " + str(len(exceltree.optimizationsarray[i])))
    compilercellsize = len(exceltree.optimizationsarray[i]) * totaltestsnum * 2

    # print("compilercellsize = " + str(compilercellsize))
    # print("compilerstartingcol[i] = " + str(exceltree.compilerstartingcol[i]))

    sheet1.cell(row=datastartrow, column=exceltree.compilerstartingcol[i]).value = exceltree.compilersarray[i]
    sheet1.merge_cells(start_row=datastartrow, start_column=exceltree.compilerstartingcol[i], end_row=datastartrow, end_column=(exceltree.compilerstartingcol[i] + compilercellsize - 1))

    ## Write optimization flags ##
    for j in range(0, len(exceltree.optimizationsarray[i])):

      optimizationcellsize = totaltestsnum * 2

      sheet1.cell(row=(datastartrow + 1), column=(exceltree.optimizationsstartingcol[i][j])).value = exceltree.optimizationsarray[i][j]
      sheet1.merge_cells(start_row=(datastartrow + 1), start_column=(exceltree.optimizationsstartingcol[i][j]), end_row=(datastartrow + 1), end_column=(exceltree.optimizationsstartingcol[i][j] + optimizationcellsize - 1))

      testcellstartingcol = exceltree.compilerstartingcol[i] + optimizationcellsize * j

      ## Write test numbers ##
      for k in range(0, len(testarray.testname)):

        # print("k = " + str(k))

        sheet1.cell(row=(datastartrow + 2), column=testcellstartingcol).value = "test_" + str(testarray.testnum[k])
        sheet1.merge_cells(start_row=(datastartrow + 2), start_column=testcellstartingcol, end_row=(datastartrow + 2), end_column=(testcellstartingcol + 1))

        testcellstartingcol = testcellstartingcol + 2

  ## Create Test iteration index ##
  sheet1.cell(row=(datastartrow + 3), column=(datastartcol - 1)).value = "Test Iteration"

  test_iteration = 1
  for i in range((datastartrow + 4), (datastartrow + 4 + testiterations)):
    sheet1.cell(row=i, column=(datastartcol - 1)).value = test_iteration

    test_iteration = test_iteration + 1

  ## Create Average row ##
  sheet1.cell(row=(datastartrow + 4 + testiterations), column=(datastartcol - 1)).value = "Average"

  ## Create Standard Deviation row ##
  sheet1.cell(row=(datastartrow + 5 + testiterations), column=(datastartcol - 1)).value = "Standard Deviation"

  ## Create File Size row ##
  sheet1.cell(row=(datastartrow + 6 + testiterations), column=(datastartcol - 1)).value = "File Size (in Bytes)"

  ## Calculate Average Execution Time and PSNR, calculate ##
  ## Standard Deviation and write Executable File Size    ##
  for dirpath, subdirnames, files in os.walk(builddir):

    for name in files:

      filesize = os.path.getsize(dirpath + '/' + name)

      ## Get test number ##
      testnum = int(name.split('_', 1)[0])

      ## Get compiler name ##
      compilername = name.rsplit('_', 1)[1].split('-')[0]
      compilerindex = exceltree.compilersarray.index(compilername)

      ## Get the optimization flags ##
      ## Split the line on the first '-' encountered and keep the right part ##
      optimizationname = name.split('-', 1)[1]
      optimizationindex = exceltree.optimizationsarray[compilerindex].index(optimizationname)

      # print(name)
      # print(testnum)
      # print(compilername)
      # print(optimizationname)

      optimizationcellsize = totaltestsnum * 2

      startingcol = exceltree.compilerstartingcol[compilerindex] + (optimizationcellsize * optimizationindex) + (testnum * 2)

      startingcell = xlref(datastartrow + 4, startingcol, False)
      endingcell = xlref(datastartrow + 4 + testiterations - 1, startingcol, False)

      ## Calculate the Average Execution time ##
      average_execution_time_cell = xlref(datastartrow + 4 + testiterations, startingcol, False)

      # print('= AVERAGE(' + startingcell + ':' + endingcell + ')')

      sheet1[average_execution_time_cell] = '= AVERAGE(' + startingcell + ':' + endingcell + ')'
      sheet1.merge_cells(start_row=(datastartrow + 4 + testiterations), start_column=startingcol, end_row=(datastartrow + 4 + testiterations), end_column=(startingcol + 1))

      ## Calculate the Standard Deviation of Execution time ##
      stdev_execution_time_cell = xlref(datastartrow + 5 + testiterations, startingcol, False)

      # print('= STDEV.P(' + startingcell + ':' + endingcell + ')')

      sheet1[stdev_execution_time_cell] = '= STDEV.P(' + startingcell + ':' + endingcell + ')'
      sheet1.merge_cells(start_row=(datastartrow + 5 + testiterations), start_column=startingcol, end_row=(datastartrow + 5 + testiterations), end_column=(startingcol + 1))

      ## Write File Size ##
      sheet1.cell(row=(datastartrow + 6 + testiterations), column=startingcol).value = filesize
      sheet1.merge_cells(start_row=(datastartrow + 6 + testiterations), start_column=startingcol, end_row=(datastartrow + 6 + testiterations), end_column=(startingcol + 1))

  print(messagecolours.INFO + "INFO: Done!" + messagecolours.DEFAULT)


  ## Add legend containing the correspondence between test number and test name ##
  legendstartrow = datastartrow + 3
  legendstartcol = 1
  # legendstartcol = datastartcol - 4

  # print("legendstartrow = " + str(legendstartrow) + ", legendstartcol = " + str(legendstartcol))

  sheet1.cell(row=legendstartrow, column=legendstartcol).value = "Legend"
  sheet1.merge_cells(start_row=legendstartrow, start_column=legendstartcol, end_row=legendstartrow, end_column=(legendstartcol + 1))

  for i in range(0, (len(testarray.testname))):
    sheet1.cell(row=(legendstartrow + i + 1), column=legendstartcol).value = "test_" + str(testarray.testnum[i])
    sheet1.cell(row=(legendstartrow + i + 1), column=(legendstartcol + 1)).value = testarray.testname[i]

  wb.save(file_path)


## ** main ** ##
def main(argv):
  syntax = "python <script_name> <source_dir> <build_dir> <output_dir> <test_iterations>"

  global testarray
  global totaltestsnum
  global datastartrow
  global datastartcol
  global testiterations

  ## Our table starts at the 6th row on column D ##
  ## pandas uses 0-based indexing while openpyxl uses 1-based indexing ##
  datastartrow = 6
  datastartcol = 10

  testiterations = 20

  print_header()

  argvlength = len(argv)

  if (argvlength == 5):
    testiterations = int(argv[4])
  elif (argvlength != 4):
    print(messagecolours.FAIL + "ERROR! Wrong number of arguments! Arguments number: " + str(argvlength) + messagecolours.DEFAULT)
    print("Re-run with: " + syntax)
    sys.exit(-2)


  testarray, totaltestsnum = count_executables(argv[1])

  exceltree = create_exceltree(argv[3])

  dump_exceltree_info(exceltree)

  if os.path.isfile('results.xlsx'):
    print("\nINFO: Removing prior results.xlsx file...")
    os.remove('results.xlsx')
    print(messagecolours.INFO + "INFO: Done!" + messagecolours.DEFAULT)


  print(messagecolours.INFO + "INFO: Creating Excel spreadsheet..." + messagecolours.DEFAULT)

  build_excel_spreadsheet(exceltree, argv[2], argv[3])

  print(messagecolours.INFO + "INFO: Finished! Exiting..." + messagecolours.DEFAULT)





## Count test files and create a Legend tables ##

## Tokenize the folder names in the output/ directory to the last '_' character ##
## to get the compiler and optimization level. Create a token array with        ##
## optimization levels and a token array with the compilers                     ##


## Build the excel spreadsheet starting from column D and line 6 ##

## On cell D9 Write "Test Iteration"; From cells D10 to D30 iterate from 1 to 20 ##

## Starting from cell E9 and E10 write "Exec. Time" "PSNR" ##
## for as many tests there are in the test directory       ##

## Starting from cell E8 merge and center pairs of cells and write ##
## "test##" for as many tests as there are in the test directory    ##

## Starting from cell E7 merge and center testnum cells and write the ##
## compiler optimization level for as many compiler optimizations     ##
## exist in each compileroptimizationarray                            ##

## Starting from cell E6 merge and (testnum * compileroptsnum) cells and write ##
## the "compilername" for as many compilers there are in the compilersarray    ##


## Read every .runlog file in the output/ directory and erase the ##
## line with the min and max result with regard to execution time ##

## From the file name find in which column we ##
## should enter the data from the CSV file    ##
## NOTE: We should include the column number for each test in the lookup tuple


## On cell D31 Write "Average Exec. Time"
## On cell D32 Write "Stardard Deviation"
## On cell D2 Write "File Size"

if (__name__ == "__main__"):
  main(sys.argv[0:])